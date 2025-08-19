import openpyxl

class LoginData:

    @staticmethod
    def get_test_data(data_num):
        book = openpyxl.load_workbook(r"C:\Users\gpnam\Downloads\RefundCookieData\TestData.xlsx")
        sheet = book.active
        result = {}

        for i in range(2, sheet.max_row + 1):
            if str(sheet.cell(row=i, column=1).value) == str(data_num):
                for j in range(2, sheet.max_column + 1):
                    key = sheet.cell(row=1, column=j).value
                    value = sheet.cell(row=i, column=j).value
                    result[key] = value

        return[result]